"""
excel_generator.py — Teklif Mukayese Excel Uretici
===================================================
Kullanim:
    python excel_generator.py "KLASOR_YOLU/veri.py"

veri.py: Sadece data iceren dosya — bkz. veri_sablonu.py
Claude bu scripte DOKUNMAZ; sadece veri.py dosyasini yazar.
"""

import sys
import importlib.util
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Windows terminal encoding
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# ── Veri dosyasini yukle ──────────────────────────────────────────────────────
if len(sys.argv) < 2:
    print("Kullanim: python excel_generator.py VERI_DOSYASI.py")
    sys.exit(1)

_spec = importlib.util.spec_from_file_location("veri", sys.argv[1])
_veri = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_veri)

CIKTI      = _veri.CIKTI
KUR_TARIHI = _veri.KUR_TARIHI
RFQ_ADI    = _veri.RFQ_ADI
KUR        = _veri.KUR
rfq_items  = _veri.rfq_items
suppliers  = _veri.suppliers
NOTLAR    = list(getattr(_veri, 'NOTLAR',    []))
AI_ANALIZ = getattr(_veri, 'AI_ANALIZ', '')

# Para birimi notlarini otomatik ekle (basa)
for _sup in reversed(suppliers):
    _curr = _sup.get('currency', 'USD')
    if _curr != 'USD':
        _rate = KUR.get(_curr, 1.0)
        NOTLAR.insert(0,
            f"\u2139 {_sup['name']} \u2013 Teklif orijinalinde {_curr} cinsinden verilmistir; "
            f"1 {_curr} = {_rate:.4f} USD ({KUR_TARIHI}) kuru ile USD'ye cevirilmistir."
        )

# ── Yardimci fonksiyonlar ─────────────────────────────────────────────────────
def fl(h):
    return PatternFill(fill_type="solid", fgColor=h)

def fn(bold=False, color="000000", size=9, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)

def al(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# Sabit stiller
HEADER_F = fl("1F3864");  HEADER_FN = fn(bold=True, color="FFFFFF", size=10)
GT_F     = fl("1F3864");  GT_FN     = fn(bold=True, color="FFFFFF", size=10)
INFO_F   = fl("2F4F4F");  INFO_FN   = fn(italic=True, color="FFFFFF", size=9)
NA_F     = fl("D3D3D3");  NA_FN     = fn(color="808080", size=9)
GREEN_F  = fl("C6EFCE");  RED_F     = fl("FFC7CE")
NUM      = "#,##0.00"
DBL_TOP  = Border(top=Side(style="double"))

FIXED = 4   # A=Item  B=Spec  C=Qty  D=Unit
SPW   = 3   # sutun sayisi/tedarikci: Birim(USD) | Toplam(USD) | Teslim

def cs(i):
    """i. tedarikcinin baslangic sutunu (1-tabanli)"""
    return FIXED + 1 + i * SPW

# ── Workbook ──────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Mukayese"
ws.row_dimensions[1].height = 38
ws.row_dimensions[2].height = 28

# ── On hesaplamalar ───────────────────────────────────────────────────────────
total_cols = FIXED + len(suppliers) * SPW

# ── Satir 1: GYP + Firma grup basliklar ──────────────────────────────────────
# NOT: MergedCell border sinirlamasi nedeniyle her section'in son sutunu
# merge'e dahil edilmez; regular cell olarak birakilir (ayni renk, border calisiyor).
ws.merge_cells(f"A1:{get_column_letter(FIXED-1)}1")
c = ws["A1"]
c.value = f"RFQ -- {RFQ_ADI}\nGuney Yildizi Petrol (GYP)"
c.fill = HEADER_F; c.font = HEADER_FN; c.alignment = al(wrap=True)
ws.cell(1, FIXED).fill = HEADER_F   # D1: regular cell → right=medium border

for i, sup in enumerate(suppliers):
    sc = cs(i)
    sl = get_column_letter(sc)
    is_last = (i == len(suppliers) - 1)
    merge_end = sc + SPW - 2 if is_last else sc + SPW - 1
    el = get_column_letter(merge_end)
    ws.merge_cells(f"{sl}1:{el}1")
    c = ws[f"{sl}1"]
    c.value = f"{sup['name']}  |  {sup['full_name']}"
    c.fill = fl(sup["color"])
    c.font = fn(bold=True, color="1F3864", size=10)
    c.alignment = al(wrap=True)
    if is_last:
        ws.cell(1, total_cols).fill = fl(sup["color"])  # son sutun: regular cell

# ── Satir 2: Alt basliklar ────────────────────────────────────────────────────
for col, h in enumerate(["Item", "Specification / Description", "Qty (RFQ)", "Unit"], 1):
    c = ws.cell(row=2, column=col, value=h)
    c.fill = HEADER_F; c.font = fn(bold=True, color="FFFFFF", size=9); c.alignment = al(wrap=True)

for i, sup in enumerate(suppliers):
    sc = cs(i)
    for j, h in enumerate(["Unit Price\n(USD)", "Total\n(USD)", "Lead\nTime"]):
        c = ws.cell(row=2, column=sc + j, value=h)
        c.fill = fl(sup["color"])
        c.font = fn(bold=True, color="1F3864", size=9)
        c.alignment = al(wrap=True)

# ── Veri satirlari ────────────────────────────────────────────────────────────
DS = 3  # data start row

for ri, item in enumerate(rfq_items):
    r = DS + ri
    ws.row_dimensions[r].height = 18
    n = item["item"]; q = item["qty"]

    ws.cell(r, 1, n).alignment = al()
    ws.cell(r, 2, item["spec"]).alignment = al(h="left", wrap=True)
    ws.cell(r, 3, q).alignment = al()
    ws.cell(r, 4, item["unit"]).alignment = al()

    valid = {}
    for sup in suppliers:
        p = sup["prices"].get(n)
        if p is not None:
            curr = sup.get("currency", "USD")
            valid[sup["name"]] = p * KUR.get(curr, 1.0)

    min_p = min(valid.values()) if valid else None
    max_p = max(valid.values()) if len(valid) > 1 else None

    for i, sup in enumerate(suppliers):
        sc = cs(i)
        p    = sup["prices"].get(n)
        curr = sup.get("currency", "USD")
        sf   = fl(sup["color"])

        p_usd = (round(p * KUR.get(curr, 1.0), 4) if curr != "USD" else p) if p is not None else None

        if p_usd is None:
            for j in range(3):
                c = ws.cell(r, sc + j, "N/A")
                c.fill = NA_F; c.font = NA_FN; c.alignment = al()
        else:
            uc = get_column_letter(sc)
            _dt  = sup.get("delivery_times", {}).get(n, sup["delivery"])
            c = ws.cell(r, sc,     p_usd);               c.fill=sf; c.font=fn(size=9); c.number_format=NUM; c.alignment=al(h="right")
            c = ws.cell(r, sc + 1, f"={uc}{r}*C{r}");   c.fill=sf; c.font=fn(size=9); c.number_format=NUM; c.alignment=al(h="right")
            c = ws.cell(r, sc + 2, _dt);                 c.fill=sf; c.font=fn(size=9); c.alignment=al()

            unit_cell = ws.cell(r, sc)
            if   min_p is not None and p_usd == min_p: unit_cell.fill = GREEN_F
            elif max_p is not None and p_usd == max_p: unit_cell.fill = RED_F

# ── Grand Total ───────────────────────────────────────────────────────────────
last = DS + len(rfq_items) - 1
gt   = last + 1
ws.row_dimensions[gt].height = 22

ws.merge_cells(f"A{gt}:{get_column_letter(FIXED-1)}{gt}")
c = ws[f"A{gt}"]
c.value = "GRAND TOTAL (USD)"; c.fill=GT_F; c.font=GT_FN; c.alignment=al(wrap=True); c.border=DBL_TOP
ws.cell(gt, FIXED).fill = GT_F   # D{gt}: regular cell → right=medium border

grand = {}
for i, sup in enumerate(suppliers):
    sc   = cs(i)
    curr = sup.get("currency", "USD")
    grand[sup["name"]] = sum(
        sup["prices"][item["item"]] * item["qty"] * KUR.get(curr, 1.0)
        for item in rfq_items
        if sup["prices"].get(item["item"]) is not None
    )
    tc = get_column_letter(sc + 1)
    dr = f"{tc}{DS}:{tc}{last}"
    for j in range(3):
        c = ws.cell(gt, sc + j); c.fill=GT_F; c.border=DBL_TOP
        if j == 1:
            c.value = f"=SUM({dr})"
            c.font=GT_FN; c.number_format=NUM; c.alignment=al(h="right")

# ── Ticari bilgi satirlari ────────────────────────────────────────────────────
for k, (label, key) in enumerate([
    ("Payment Terms",    "payment"),
    ("Incoterm",         "incoterm"),
    ("Delivery Location","location"),
]):
    r2 = gt + 1 + k
    ws.row_dimensions[r2].height = 16
    ws.merge_cells(f"A{r2}:{get_column_letter(FIXED-1)}{r2}")
    c = ws[f"A{r2}"]; c.value=label; c.fill=INFO_F; c.font=INFO_FN; c.alignment=al()
    ws.cell(r2, FIXED).fill = INFO_F   # D{r2}: regular cell → right=medium border
    for i, sup in enumerate(suppliers):
        sc = cs(i)
        is_last = (i == len(suppliers) - 1)
        merge_end = sc + SPW - 2 if is_last else sc + SPW - 1
        el = get_column_letter(merge_end)
        ws.merge_cells(f"{get_column_letter(sc)}{r2}:{el}{r2}")
        c = ws[f"{get_column_letter(sc)}{r2}"]
        c.value = sup.get(key, "Belirtilmemis"); c.fill=INFO_F; c.font=INFO_FN; c.alignment=al()
        if is_last:
            ws.cell(r2, total_cols).fill = INFO_F  # son sutun: regular cell

# ── Dipnot ────────────────────────────────────────────────────────────────────
nr = gt + 5
ws.row_dimensions[nr].height = 24
ws.merge_cells(f"A{nr}:{get_column_letter(total_cols)}{nr}")
c = ws[f"A{nr}"]
c.value = (
    f"Kur tarihi: {KUR_TARIHI}  |  "
    "Tum teklifler USD bazinda karsilastirilmistir.  |  "
    "Tablodaki toplam degerler RFQ miktarlari uzerinden hesaplanmistir."
)
c.font = fn(italic=True, color="7F7F7F", size=8); c.alignment = al(h="left", wrap=True)

# ── Sutun genislikleri ────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 55
ws.column_dimensions["C"].width = 8
ws.column_dimensions["D"].width = 6
for i in range(len(suppliers)):
    sc = cs(i)
    for offset, width in [(0, 14), (1, 14), (2, 14)]:
        ws.column_dimensions[get_column_letter(sc + offset)].width = width

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:{get_column_letter(total_cols)}2"

# ── Cizgiler (Borders) ────────────────────────────────────────────────────────
# openpyxl MergedCell sinirlamasi: non-topleft hucrelere border yazilmiyor.
# Cozum: her section'in son sutunu merge disinda regular cell birakildi (yukarda).
# Boylece thick/medium borderlar tum satirlarda tutarli cikiyor.
thin_s   = Side(style="thin")
medium_s = Side(style="medium")
thick_s  = Side(style="thick")

def apply_border(cell, left=None, right=None, top=None, bottom=None):
    """MergedCell'leri atlayarak border uygular (AttributeError: pass)."""
    try:
        b = cell.border
        cell.border = Border(
            left   = left   if left   is not None else b.left,
            right  = right  if right  is not None else b.right,
            top    = top    if top    is not None else b.top,
            bottom = bottom if bottom is not None else b.bottom,
        )
    except AttributeError:
        pass

last_info  = gt + 3
table_rows = range(1, last_info + 1)
all_cols_r = range(1, total_cols + 1)

# 1. Tum tabloya ince cizgi
for r in table_rows:
    for col in all_cols_r:
        apply_border(ws.cell(r, col), thin_s, thin_s, thin_s, thin_s)

# 2. Satir 1 alt kenari → orta
for col in all_cols_r:
    apply_border(ws.cell(1, col), bottom=medium_s)

# 3. Satir 2 alt kenari → orta
for col in all_cols_r:
    apply_border(ws.cell(2, col), bottom=medium_s)

# 4. Sabit bolum sag kenari (D sutunu) → orta
for r in table_rows:
    apply_border(ws.cell(r, FIXED), right=medium_s)

# 5. Her tedarikcinin sol kenari → orta
for i in range(len(suppliers)):
    sc_b = cs(i)
    for r in table_rows:
        apply_border(ws.cell(r, sc_b), left=medium_s)

# 6. Grand Total ust kenari → orta
for col in all_cols_r:
    apply_border(ws.cell(gt, col), top=medium_s)

# 7. Dis cerceve → kalin
for col in all_cols_r:
    apply_border(ws.cell(1, col), top=thick_s)
for col in all_cols_r:
    apply_border(ws.cell(last_info, col), bottom=thick_s)
for r in table_rows:
    apply_border(ws.cell(r, 1), left=thick_s)
for r in table_rows:
    apply_border(ws.cell(r, total_cols), right=thick_s)

# ── Notlar & AI Analizi ───────────────────────────────────────────────────────
# Sinir cizgisi yok; baslik satiri + kalem satırları sade renk bloklari.
note_row = nr + 2

if NOTLAR:
    ws.merge_cells(f"A{note_row}:{get_column_letter(total_cols)}{note_row}")
    c = ws[f"A{note_row}"]
    c.value = "ANOMALILER & NOTLAR"
    c.fill = fl("F4B942"); c.font = fn(bold=True, color="FFFFFF", size=9)
    c.alignment = al(h="left")
    ws.row_dimensions[note_row].height = 18
    note_row += 1

    for note in NOTLAR:
        ws.merge_cells(f"A{note_row}:{get_column_letter(total_cols)}{note_row}")
        c = ws[f"A{note_row}"]
        c.value = note
        c.fill = fl("FFF2CC"); c.font = fn(color="000000", size=9)
        c.alignment = al(h="left", wrap=True)
        ws.row_dimensions[note_row].height = 20
        note_row += 1

    note_row += 1  # bos satir

# AI_ANALIZ: list veya str her ikisi desteklenir
if isinstance(AI_ANALIZ, str):
    ai_items = [s.strip() for s in AI_ANALIZ.split('\n') if s.strip()] if AI_ANALIZ else []
else:
    ai_items = [x for x in AI_ANALIZ if x]

if ai_items:
    ws.merge_cells(f"A{note_row}:{get_column_letter(total_cols)}{note_row}")
    c = ws[f"A{note_row}"]
    c.value = "YAPAY ZEKA ANALIZI"
    c.fill = fl("2E75B6"); c.font = fn(bold=True, color="FFFFFF", size=9)
    c.alignment = al(h="left")
    ws.row_dimensions[note_row].height = 18
    note_row += 1

    for item in ai_items:
        ws.merge_cells(f"A{note_row}:{get_column_letter(total_cols)}{note_row}")
        c = ws[f"A{note_row}"]
        c.value = f"\u2022  {item}"
        c.fill = fl("DEEAF1"); c.font = fn(italic=True, color="000000", size=9)
        c.alignment = al(h="left", wrap=True)
        ws.row_dimensions[note_row].height = 20
        note_row += 1

# ── Sheet 2: Kur Bilgisi ──────────────────────────────────────────────────────
ws2 = wb.create_sheet("Kur Bilgisi")
ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 38

kur_satirlari = [
    ("Kur Tarihi", KUR_TARIHI),
    ("Kaynak",     "exchangerate-api.com"),
    ("", ""),
    ("Para Birimi", "1 Birim = ? USD"),
    ("USD", "1.0000"),
]
for cur, rate in KUR.items():
    if cur != "USD":
        kur_satirlari.append((cur, f"{rate:.4f}  (1 USD = {1/rate:.4f} {cur})"))
kur_satirlari += [("", ""), ("Not", "Tum teklifler USD bazinda karsilastirildi.")]

for r2, (k, v) in enumerate(kur_satirlari, 1):
    ws2.cell(r2, 1, k).font = fn(bold=True)
    ws2.cell(r2, 2, v)

# ── Sheet 3: Ham Veri ─────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Ham Veri")
hdrs = ["Tedarikci", "Odeme", "Incoterm", "Konum",
        "Hesaplanan Toplam (RFQ Qty, USD)", "Notlar"]
for col, h in enumerate(hdrs, 1):
    ws3.cell(1, col, h).font = fn(bold=True)
    ws3.column_dimensions[get_column_letter(col)].width = 30

for i, sup in enumerate(suppliers, 2):
    ws3.cell(i, 1, sup["full_name"])
    ws3.cell(i, 2, sup["payment"])
    ws3.cell(i, 3, sup["incoterm"])
    ws3.cell(i, 4, sup["location"])
    c = ws3.cell(i, 5, grand[sup["name"]])
    c.number_format = NUM

# ── Kaydet ────────────────────────────────────────────────────────────────────
wb.save(CIKTI)

print(f"\nTAMAM: {CIKTI}")
print(f"\n{'-'*50}")
print(f"{'GRAND TOTAL (USD, RFQ Miktarlarinda)':^50}")
print(f"{'-'*50}")
for name, total in grand.items():
    print(f"  {name:<20} ${total:>12,.2f}")
print(f"{'-'*50}")
if grand:
    mn = min(grand, key=grand.get)
    mx = max(grand, key=grand.get)
    print(f"  En dusuk  --> {mn}: ${grand[mn]:,.2f}")
    print(f"  En yuksek --> {mx}: ${grand[mx]:,.2f}")
